import openpyxl as xl

def check_mail(email):
    row_maximum = ws1.max_row

    for i in range(2, row_maximum + 1):

        check = ws2.cell(row=i, column=2).value
        if check == email:
            return (True, i)

    return (False, 0)

filename = "Final.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]

filename = "Final_3.xlsx"
wb2 = xl.load_workbook(filename)
ws2 = wb2.worksheets[0]


mr = ws2.max_row
mc = ws2.max_column

n = ws1.max_row+1

for i in range(2, mr + 1):
    flag = True
    email = ws2.cell(row=i,column=2).value
    temp = check_mail(email)
    flag = temp[0]
    row_num = temp[1]

    if flag :
        a = ws2.cell(row=i, column=3)
        ws1.cell(row=row_num, column=5).value = a.value
    else:
        a=ws2.cell(row=i,column=1)
        b=ws2.cell(row=i,column=2)
        c=ws2.cell(row=i,column=3)

        ws1.cell(row=n,column=1).value= a.value
        ws1.cell(row=n,column=2).value=b.value
        ws1.cell(row=n,column=3).value=0
        ws1.cell(row=n,column=4).value=0
        ws1.cell(row=n,column=5).value=c.value

        n+=1



wb1.save(str("Final.xlsx"))