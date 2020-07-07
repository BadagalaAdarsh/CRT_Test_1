import openpyxl as xl


# def check_mail_in_final(mail):
#     row_maximum = ws3.max_row
#
#     for i in range(2, row_maximum + 1):
#         check = ws3.cell(row=i, column=2).value
#
#         if check == email:
#             return (True, i)
#
#     return (False, 0)


def check_mail_one(email):
    row_maximum = ws2.max_row

    for i in range(2, row_maximum + 1):

        check = ws2.cell(row=i, column=3).value
        if check == email:
            return (True, i)

    return (False, 0)


# def check_mail_two(email):
#     row_maximum = ws5.max_row
#
#     for i in range(2, row_maximum + 1):
#
#         check = ws5.cell(row=i, column=3).value
#         if check == email:
#             return (True, i)
#
#     return (False, 0)


def check_plagarism_one(row_val, col_val):
    # column value is 12 for first question result
    var = ws2.cell(row=row_val, column=5).value

    if var == "" or var == " " or var == "-":
        return False

    try:
        var = var[:5]
        if float(var) > 60:
            return True
        else:
            return False
    except:
        return False


# def check_plagarism_two(row_val, col_val):
#     # column value is 12 for first question result
#     var = ws5.cell(row=row_val, column=5).value
#
#     if var == "" or var == " " or var == "-":
#         return False
#
#     try:
#         var = var[:5]
#         if float(var) >= 75:
#             return True
#         else:
#             return False
#     except:
#         return False


filename = "Q2_Result.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]

filename = "Q2_Plag.xlsx"
wb2 = xl.load_workbook(filename)
ws2 = wb2.worksheets[0]

filename = 'Final_2.xlsx'
wb3 = xl.load_workbook(filename)
ws3 = wb3.worksheets[0]


mr = ws1.max_row
mc = ws1.max_column

for i in range(1, mr + 1):
    flag = True
    for j in range(1, mc + 1):

        if j == 2:
            email = ws1.cell(row=i, column=j).value
            temp = check_mail_one(email)
            flag = temp[0]
            row = temp[1]

        if j >= 3 and j <= 10:
            continue

        if flag:
            if j == 11:

                isplag = check_plagarism_one(row, j)

                if isplag:
                    ws3.cell(row=i, column=3).value = 0

                else:
                    c = ws1.cell(row=i, column=j)
                    ws3.cell(row=i, column=3).value = c.value

                continue

        c = ws1.cell(row=i, column=j)
        if j == 11:
            ws3.cell(row=i, column=3).value = c.value
        else:
            ws3.cell(row=i, column=j).value = c.value

wb3.save(str("Final_2.xlsx"))


