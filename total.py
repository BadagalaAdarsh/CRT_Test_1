import openpyxl as xl

filename = 'Final.xlsx'
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]

mr = ws1.max_row
mc = ws1.max_column

for i in range (2,mr+1):
        
    a = ws1.cell(row=i, column=3).value
    b =ws1.cell(row=i, column=4).value
    c = ws1.cell(row=i, column=5).value

    ws1.cell(row=i,column=6).value = int(a)+int(b)+int(c)

wb1.save(str("Final.xlsx"))